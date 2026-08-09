from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import config

def create_excel_report(scheduler, solver):

    filename = "TennisScheduler_Result.xlsx"

    wb = Workbook()

    # ==================================================
    # RIMOZIONE FOGLIO INIZIALE
    # ==================================================

    ws = wb.active
    wb.remove(ws)

    # ==================================================
    # STILI
    # ==================================================

    title_font = Font(
        size=18,
        bold=True
    )

    subtitle_font = Font(
        size=11,
        italic=True
    )

    header_font = Font(
        bold=True
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    ok_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    warning_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC"
    )

    alternate_fill = PatternFill(
        fill_type="solid",
        fgColor="F7F7F7"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ==================================================
    # FOGLIO 1 RIEPILOGO
    # ==================================================

    ws = wb.create_sheet("Riepilogo")

    ws["A1"] = "TENNIS SCHEDULER"
    ws["A1"].font = title_font

    ws["A2"] = "Risultato torneo"
    ws["A2"].font = subtitle_font

    ws["A4"] = "Voce"
    ws["B4"] = "Risultato"

    for cell in ws[4]:

        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    selected_matches = sum(
        solver.Value(var)
        for var in scheduler.match_vars.values()
    )

    duplicate_matches = 0

    men_deviation = sum(
        solver.Value(var)
        for var in scheduler.men_deviation_vars.values()
    )

    soft_avoid = sum(
        solver.Value(var)
        for var in scheduler.soft_avoid_vars
    )

    different_opponents = sum(
        solver.Value(var)
        for var in scheduler.opponent_played_vars.values()
    )

    summary = [
        (
            "Partite selezionate",
            f"{selected_matches} / {config.NUM_MATCHES}"
        ),
        (
            "Partite duplicate",
            duplicate_matches
        ),
        (
            "Deviazione uomini",
            men_deviation
        ),
        (
            "Incontri indesiderati",
            soft_avoid
        ),
        (
            "Avversari diversi",
            different_opponents
        ),
        (
            "Verifica finale",
            "OK"
        ),
    ]

    row = 5

    for label, value in summary:

        ws.cell(
            row=row,
            column=1
        ).value = label

        ws.cell(
            row=row,
            column=2
        ).value = value

        ws.cell(
            row=row,
            column=1
        ).border = thin_border

        ws.cell(
            row=row,
            column=2
        ).border = thin_border

        if label == "Verifica finale":

            ws.cell(
                row=row,
                column=2
            ).font = Font(bold=True)

            ws.cell(
                row=row,
                column=2
            ).fill = ok_fill

            ws.cell(
                row=row,
                column=2
            ).alignment = Alignment(
                horizontal="center"
            )

        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    ws.sheet_view.showGridLines = False

    # ==================================================
    # FOGLIO 2 PARTITE
    # ==================================================

    ws = wb.create_sheet("Partite")

    headers = [
      "N.",
      "Coppia 1",
      "Coppia 2"
    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=1,
            column=col
        )

        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    row = 2
    match_number = 1

    for i, match in enumerate(scheduler.matches):

        if solver.Value(scheduler.match_vars[i]):

            ws.cell(
                row=row,
                column=1
            ).value = match_number

            ws.cell(
                row=row,
                column=2
            ).value = str(match.pair1)

            ws.cell(
                row=row,
                column=3
            ).value = str(match.pair2)

            for col in range(1, 4):

                cell = ws.cell(
                    row=row,
                    column=col
                )

                cell.border = thin_border

            ws.cell(
                row=row,
                column=4
            ).alignment = Alignment(
                horizontal="center"
            )

            row += 1
            match_number += 1

    last_row = row - 1

    if last_row >= 2:

        table = Table(
            displayName="MatchesTable",
            ref=f"A1:C{last_row}"
          )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style

        ws.add_table(table)

    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28

    ws.sheet_view.showGridLines = False

    # ==================================================
    # FOGLIO 3 GIOCATORI
    # ==================================================

    ws = wb.create_sheet("Giocatori")


    # ==================================================
    # PARTITE PER COPPIA
    # ==================================================

    pair_match_counts = {}

    for i, match in enumerate(scheduler.matches):

        if solver.Value(scheduler.match_vars[i]):

            pair1 = (
                match.pair1.man,
                match.pair1.woman
            )

            pair2 = (
                match.pair2.man,
                match.pair2.woman
            )

            pair_match_counts[pair1] = (
                pair_match_counts.get(pair1, 0) + 1
            )

            pair_match_counts[pair2] = (
                pair_match_counts.get(pair2, 0) + 1
            )


    # ==================================================
    # COPPIE
    # ==================================================

    ws["A1"] = "COPPIE"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].border = thin_border


    headers = [
        "Coppia",
        "Partite",
    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=2,
            column=col
        )

        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    row = 3

    for man, woman in config.PAIRS:

        pair = (man, woman)

        ws.cell(
            row=row,
            column=1
        ).value = f"{man} - {woman}"

        ws.cell(
            row=row,
            column=2
        ).value = pair_match_counts.get(pair, 0)

        for col in range(1, 3):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.border = thin_border

            if row % 2 == 1:

                cell.fill = alternate_fill

        ws.cell(
            row=row,
            column=2
        ).alignment = Alignment(
            horizontal="center"
        )

        row += 1


    last_pair_row = row - 1


    # ==================================================
    # GIOCATORI
    # ==================================================

    row += 2

    ws.cell(
        row=row,
        column=1
    ).value = "GIOCATORI"

    ws.cell(
        row=row,
        column=1
    ).font = header_font

    ws.cell(
        row=row,
        column=1
    ).fill = header_fill

    ws.cell(
        row=row,
        column=1
    ).border = thin_border


    row += 1


    headers = [
        "Giocatore",
        "Sesso",
        "Partite",
        "Target",
        "Deviazione",
    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=row,
            column=col
        )

        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    row += 1

    players_start_row = row

    for player in sorted(scheduler.player_count_vars):

        count = solver.Value(
            scheduler.player_count_vars[player]
        )

        if player in config.MEN:

            gender = "M"

            target = scheduler.target_men_matches

            deviation = solver.Value(
                scheduler.men_deviation_vars[player]
            )

        else:

            gender = "F"

            target = scheduler.target_women_matches

            deviation = 0

        values = [
            player,
            gender,
            count,
            target,
            deviation,
        ]

        for col, value in enumerate(values, start=1):

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.value = value
            cell.border = thin_border

            if row % 2 == 0:

                cell.fill = alternate_fill

        if deviation != 0:

            ws.cell(
                row=row,
                column=5
            ).fill = warning_fill

        for col in range(2, 6):

            ws.cell(
                row=row,
                column=col
            ).alignment = Alignment(
                horizontal="center"
            )

        row += 1


    last_player_row = row - 1


    # ==================================================
    # TABELLA COPPIE
    # ==================================================

    if last_pair_row >= 3:

        table = Table(
            displayName="PairsTable",
            ref=f"A2:B{last_pair_row}"
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style

        ws.add_table(table)


    # ==================================================
    # TABELLA GIOCATORI
    # ==================================================

    if last_player_row >= players_start_row:

        table = Table(
            displayName="PlayersTable",
            ref=f"A{players_start_row - 1}:E{last_player_row}"
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style

        ws.add_table(table)


    # ==================================================
    # FORMATO
    # ==================================================

    ws.freeze_panes = "A3"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    ws.sheet_view.showGridLines = False

    # ==================================================
    # FOGLIO 4 AVVERSARI
    # ==================================================

    ws = wb.create_sheet("Avversari")

    # --------------------------------------------------
    # Funzione per ottenere il numero di incontri
    # tra due giocatori
    # --------------------------------------------------

    def get_opponent_count(player1, player2):

        key = tuple(sorted((player1, player2)))

        count_var = scheduler.opponent_count_vars.get(key)

        if count_var is None:
            return 0

        return solver.Value(count_var)


    # --------------------------------------------------
    # Stile celle triangolo non utilizzato
    # --------------------------------------------------

    hidden_fill = PatternFill(
        fill_type="solid",
        fgColor="D9D9D9"
    )


    # ==================================================
    # 1. DONNE × DONNE
    # ==================================================

    women = sorted(config.WOMEN)

    start_row = 1

    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=len(women) + 1
    )

    cell = ws.cell(
        row=start_row,
        column=1
    )

    cell.value = "DONNE × DONNE"
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    header_row = start_row + 1

    # intestazione colonne
    ws.cell(
        row=header_row,
        column=1
    ).value = ""

    for j, woman in enumerate(women, start=2):

        cell = ws.cell(
            row=header_row,
            column=j
        )

        cell.value = woman
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # righe della matrice
    for i, woman in enumerate(women):

        row = header_row + 1 + i

        cell = ws.cell(
            row=row,
            column=1
        )

        cell.value = woman
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        for j in range(len(women)):

            col = j + 2

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            # diagonale
            if i == j:

                cell.value = "—"
                cell.fill = header_fill

            # triangolo superiore
            elif j > i:

                cell.value = get_opponent_count(
                    woman,
                    women[j]
                )

                if row % 2 == 0:
                    cell.fill = alternate_fill

            # triangolo inferiore: oscurato
            else:

                cell.value = ""
                cell.fill = hidden_fill


    # ==================================================
    # 2. UOMINI × UOMINI
    # ==================================================

    men = sorted(config.MEN)

    start_row = header_row + len(women) + 3

    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=len(men) + 1
    )

    cell = ws.cell(
        row=start_row,
        column=1
    )

    cell.value = "UOMINI × UOMINI"
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    header_row = start_row + 1

    # intestazione colonne
    ws.cell(
        row=header_row,
        column=1
    ).value = ""

    for j, man in enumerate(men, start=2):

        cell = ws.cell(
            row=header_row,
            column=j
        )

        cell.value = man
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # righe della matrice
    for i, man in enumerate(men):

        row = header_row + 1 + i

        cell = ws.cell(
            row=row,
            column=1
        )

        cell.value = man
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        for j in range(len(men)):

            col = j + 2

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            # diagonale
            if i == j:

                cell.value = "—"
                cell.fill = header_fill

            # triangolo superiore
            elif j > i:

                cell.value = get_opponent_count(
                    man,
                    men[j]
                )

                if row % 2 == 0:
                    cell.fill = alternate_fill

            # triangolo inferiore: oscurato
            else:

                cell.value = ""
                cell.fill = hidden_fill


    # ==================================================
    # 3. UOMINI × DONNE
    # ==================================================

    start_row = header_row + len(men) + 3

    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=len(women) + 1
    )

    cell = ws.cell(
        row=start_row,
        column=1
    )

    cell.value = "UOMINI × DONNE"
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    header_row = start_row + 1

    # intestazione colonne
    ws.cell(
        row=header_row,
        column=1
    ).value = ""

    for j, woman in enumerate(women, start=2):

        cell = ws.cell(
            row=header_row,
            column=j
        )

        cell.value = woman
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # righe della matrice
    for i, man in enumerate(men):

        row = header_row + 1 + i

        cell = ws.cell(
            row=row,
            column=1
        )

        cell.value = man
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        for j, woman in enumerate(women):

            col = j + 2

            cell = ws.cell(
                row=row,
                column=col
            )

            cell.value = get_opponent_count(
                man,
                woman
            )

            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            if row % 2 == 0:
                cell.fill = alternate_fill


    # ==================================================
    # IMPOSTAZIONI FOGLIO
    # ==================================================

    ws.freeze_panes = "B3"

    ws.column_dimensions["A"].width = 20

    for col in range(2, len(women) + 2):

        ws.column_dimensions[
            get_column_letter(col)
        ].width = 14

    ws.sheet_view.showGridLines = False


    # ==================================================
    # SALVATAGGIO
    # ==================================================

    wb.save(filename)

    print(
        f"\nFile Excel creato: {filename}"
    )

    return filename
